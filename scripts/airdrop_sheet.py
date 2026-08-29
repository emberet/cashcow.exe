#!/usr/bin/env python3
"""Render airdrop-allocation.json into a reviewable .xlsx workbook.

Three sheets: the allocation itself, everyone excluded and why, and the
parameters the run used so the numbers can be reproduced.
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(sys.argv[1] if len(sys.argv) > 1 else ".")
d = json.loads((OUT / "airdrop-allocation.json").read_text())

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
MUTED = Font(color="666666", italic=True, size=10)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
EXCL_FILL = PatternFill("solid", fgColor="FCE4E4")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_header(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------------------------------------------------------------- allocation
ws = wb.active
ws.title = "Allocation"
headers = ["#", "Wallet", "Tokens held", "Tokens 6h ago", "% of qualifying",
           "SOL", "USD", "Wallet has SOL?", "Basis"]
write_header(ws, headers)

rows = sorted(d["qualifying"], key=lambda r: -r["tokens"])
for i, r in enumerate(rows, start=1):
    vals = [
        i,
        r["owner"],
        round(r["tokens"], 6),
        round(r.get("tokensBefore") or 0, 6),
        round(r["pctOfQualifying"], 4) / 100,
        round(r["allocSol"], 9),
        round(r["allocUsd"], 2),
        "no — created by transfer" if r.get("needsRent") else "yes",
        r.get("basis") or "",
    ]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=i + 1, column=c, value=v)
        cell.border = BORDER
        if c in (3, 4):
            cell.number_format = "#,##0"
        elif c == 5:
            cell.number_format = "0.0000%"
        elif c == 6:
            cell.number_format = "0.000000000"
        elif c == 7:
            cell.number_format = '"$"#,##0.00'
        if c == 2:
            cell.font = Font(name="Menlo", size=9)
        if r.get("needsRent"):
            cell.fill = WARN_FILL

total_row = len(rows) + 2
ws.cell(row=total_row, column=2, value=f"TOTAL — {len(rows)} recipients").font = Font(bold=True)
for col, val, fmt in ((3, sum(r["tokens"] for r in rows), "#,##0"),
                      (6, d["distributed"], "0.000000000"),
                      (7, d["distributed"] * d["solUsd"], '"$"#,##0.00')):
    c = ws.cell(row=total_row, column=col, value=round(val, 9))
    c.font, c.number_format, c.border = Font(bold=True), fmt, BORDER

ws.cell(row=total_row + 2, column=2,
        value="Highlighted rows are wallets holding no SOL. The transfer creates "
              "their account; each is above the rent-exempt minimum.").font = MUTED
autosize(ws, [5, 46, 18, 18, 15, 16, 13, 22, 52])

# ---------------------------------------------------------------- excluded
we = wb.create_sheet("Excluded")
write_header(we, ["#", "Wallet", "Tokens held", "Tokens 6h ago", "Reason excluded"])
ex = sorted(d["excluded"], key=lambda r: -r["tokens"])
for i, r in enumerate(ex, start=1):
    vals = [i, r["owner"], round(r["tokens"], 6),
            round(r.get("tokensBefore") or 0, 6), r["reason"]]
    for c, v in enumerate(vals, start=1):
        cell = we.cell(row=i + 1, column=c, value=v)
        cell.border = BORDER
        if c in (3, 4):
            cell.number_format = "#,##0"
        if c == 2:
            cell.font = Font(name="Menlo", size=9)
        if "pool/curve" in r["reason"]:
            cell.fill = EXCL_FILL
we.cell(row=len(ex) + 3, column=2,
        value="Shaded rows are program accounts (AMM pool / bonding curve), not people. "
              "The pool alone would have taken ~38% of the pot.").font = MUTED
autosize(we, [5, 46, 18, 18, 74])

# ---------------------------------------------------------------- parameters
wp = wb.create_sheet("Run parameters")
wp.cell(row=1, column=1, value="Airdrop allocation — run parameters").font = TITLE_FONT
facts = [
    ("Generated at (UTC)", d["generatedAt"]),
    ("Slot", d["slot"]),
    ("Mint", d["mint"]),
    ("Sender wallet", d["sender"]),
    ("", ""),
    ("Threshold", f'{d["threshold"]:,.0f} tokens'),
    ("Holding period", f'{d["hours"]}h  (cutoff {d["cutoffIso"]})'),
    ("Rule", "held >= threshold NOW and >= threshold at the cutoff"),
    ("Allocation method", "pro-rata on tokens held across qualifying holders"),
    ("", ""),
    ("Wallet balance", f'{d["walletSol"]:.9f} SOL'),
    ("Kept back", f'{d["keep"]:.9f} SOL'),
    ("Network fee reserve", f'{d["networkFee"]:.9f} SOL'),
    ("Service fee reserve", f'{d["serviceFee"]:.9f} SOL  (slerf.tools — VERIFY on the site)'),
    ("POT DISTRIBUTED", f'{d["distributed"]:.9f} SOL'),
    ("Unallocated dust", f'{d["dust"]:.9f} SOL'),
    ("", ""),
    ("SOL price", f'${d["solUsd"]} ({d["priceSource"]})'),
    ("Total USD distributed", f'${d["distributed"] * d["solUsd"]:,.2f}'),
    ("", ""),
    ("Token supply", f'{d["totalSupply"]:,.2f}'),
    ("Owners in census", d["censusOwners"]),
    ("Candidates >= threshold", d["candidates"]),
    ("Qualifying", len(d["qualifying"])),
    ("Excluded", len(d["excluded"])),
    ("Below threshold", d["belowThreshold"]),
    ("Unfunded recipients", f'{d.get("unfundedRecipients", 0)} (rent minimum {d.get("rentMin", 0)} SOL)'),
]
for i, (k, v) in enumerate(facts, start=3):
    kc = wp.cell(row=i, column=1, value=k)
    kc.font = Font(bold=True) if k else Font()
    wp.cell(row=i, column=2, value=v)
autosize(wp, [26, 78])

path = OUT / "airdrop-allocation.xlsx"
wb.save(path)
print(f"wrote {path}")
print(f"  Allocation: {len(rows)} rows   Excluded: {len(ex)} rows")
